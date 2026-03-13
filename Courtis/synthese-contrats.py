#!/usr/bin/env python3
"""
Extraction des contrats d'assurance depuis les exports Brio (PDF)
et génération d'un fichier Excel de synthèse.
"""

import os
import re
import glob
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
PDF_DIR = os.path.join(os.path.dirname(__file__), "Synthese prod leuze ")
OUTPUT_FILE = os.path.join(PDF_DIR, "Synthese_Contrats_Groupe_Nonet.xlsx")


def extract_text_from_pdf(pdf_path):
    """Extrait le texte brut d'un PDF."""
    reader = PyPDF2.PdfReader(pdf_path)
    text = ""
    for page in reader.pages:
        t = page.extract_text()
        if t:
            text += t + "\n"
    return text


def clean_brio_headers(text):
    """Supprime les en-têtes/pieds de page Brio répétés qui collent aux données de contrats."""
    # Supprime les blocs "Votre dossier ... Numéro d'entreprise: XXXXXXX"
    # IMPORTANT: s'arrête juste après le numéro (ne mange pas le contrat qui suit)
    cleaned = re.sub(
        r"Votre dossier assurances[^\n]*\n(?:.*?Num.ro d.entreprise:\s*\d+)\s*",
        "\n", text, flags=re.DOTALL
    )
    # Nettoyer les lignes "13 mars 2026  Contrat(s)" résiduelles
    cleaned = re.sub(r"\d+ \w+ \d{4}\s+Contrat\(s\)\s*", "\n", cleaned)
    return cleaned


def extract_societe(text):
    """Extrait le nom de la société depuis la page 1."""
    # Pattern robuste (tolère variations d'accents et d'apostrophes)
    m = re.search(r"Donn.es personnelles\s*\n?(.*?)Num.ro d.entreprise", text, re.DOTALL)
    if m:
        name = m.group(1).strip()
        # Nettoyer les préfixes courants
        name = re.sub(r"^A la direction de\s+", "", name, flags=re.IGNORECASE)
        name = re.sub(r"^(Monsieur|Madame|M\.|Mme)\s+", "", name, flags=re.IGNORECASE)
        name = re.sub(r"^et (Monsieur|Madame|M\.|Mme)\s+", "", name, flags=re.IGNORECASE)
        # Supprimer lignes vides
        lines = [l.strip() for l in name.split("\n") if l.strip()]
        if lines:
            return lines[0]
    # Fallback: nom du fichier
    return None


def extract_contrats(text):
    """Parse les contrats depuis le texte Brio."""
    contrats = []

    # Pattern pour détecter le début d'un contrat:
    # "N. Type - Compagnie - N°Police - Description"
    # suivi de la ligne Etat/Echéance
    contract_pattern = re.compile(
        r"(\d+)\.\s+"  # numéro du contrat
        r"(.+?)\s*-\s*"  # type de contrat (branche)
        r"([A-Z][A-Za-zÀ-ÿ\s&'.]+?(?:SA|SRL|NV|SPRL|AG|SE|ASSISTANCE|EUROPE|PROTECT|INSURANCE)?)\s*-\s*"  # compagnie
        r"(\d[\d\s/]*\d)\s*"  # n° police
        r"(?:-\s*(.+?))?\s*\n"  # description optionnelle
    )

    # On cherche aussi Echéance et Prime après chaque contrat
    echeance_pattern = re.compile(r"Ech[ée]ance:\s*(\d{2}/\d{2})")
    prime_pattern = re.compile(r"Derni[eè]re prime annuelle:\s*€\s*([\d\s.,]+)")
    periodicite_pattern = re.compile(r"P[ée]riodicit[ée]:\s*(\w+)")

    # Découper le texte en sections par contrat
    # Approche alternative plus robuste: chercher tous les patterns "N. ... - ... - NUMERO"
    # puis extraire les infos entre deux contrats consécutifs

    # Trouver toutes les positions de début de contrat
    # Pattern souple: un numéro suivi d'un point et d'un espace, possiblement après un saut de page
    starts = []
    for m in re.finditer(r"(?:^|\n|\s)(\d+)\.\s+[A-Z]", text):
        # Vérifier que c'est bien un contrat (suivi d'un "Etat:" dans les ~500 chars suivants)
        lookahead = text[m.start():m.start() + 800]
        if re.search(r"Etat:", lookahead):
            starts.append(m.start())

    for i, start in enumerate(starts):
        end = starts[i + 1] if i + 1 < len(starts) else len(text)
        section = text[start:end]

        # Extraire la première ligne du contrat (peut être sur plusieurs lignes avant "Etat:")
        header_match = re.search(r"(\d+)\.\s+(.+?)(?=\s*Etat:)", section, re.DOTALL)
        if not header_match:
            continue

        num = header_match.group(1)
        header_raw = header_match.group(2).replace("\n", " ").strip()

        # Parser le header: Type - Compagnie - N°Police - Description
        parts = [p.strip() for p in header_raw.split(" - ")]
        if len(parts) < 3:
            continue

        branche = parts[0]  # type/branche d'assurance

        # La compagnie est généralement le 2e élément
        compagnie = parts[1]

        # Le n° police est le premier élément qui ressemble à un numéro
        n_police = ""
        description_parts = []
        for p in parts[2:]:
            if re.match(r"^\d[\d\s/]+$", p.strip()) and not n_police:
                n_police = p.strip()
            else:
                description_parts.append(p)

        description = " - ".join(description_parts) if description_parts else ""

        # Objet du risque = branche + description
        objet = branche
        if description:
            objet = f"{branche} — {description}"

        # Échéance
        ech_match = echeance_pattern.search(section)
        echeance = ech_match.group(1) if ech_match else ""

        # Périodicité
        per_match = periodicite_pattern.search(section)
        periodicite = per_match.group(1) if per_match else ""

        # Prime
        prime_match = prime_pattern.search(section)
        prime_str = ""
        prime_val = None
        if prime_match:
            prime_str = prime_match.group(1).strip()
            # Convertir en nombre: "5.788,00" -> 5788.00
            try:
                prime_val = float(prime_str.replace(" ", "").replace(".", "").replace(",", "."))
            except ValueError:
                prime_val = None

        contrats.append({
            "compagnie": compagnie,
            "objet_risque": objet,
            "n_police": n_police,
            "echeance": echeance,
            "periodicite": periodicite,
            "prime_str": prime_str,
            "prime_val": prime_val,
        })

    return contrats


def _apply_styles_to_sheet(ws, headers, col_widths):
    """Applique les styles communs à un onglet."""
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_headers(ws, headers):
    """Écrit la ligne d'en-têtes."""
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _write_subtotal_row(ws, row, label, prime_sum, nb_cols):
    """Écrit une ligne de sous-total."""
    subtotal_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    subtotal_font = Font(name="Calibri", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subtotal_fill
        cell.font = subtotal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=row, column=1, value=label)
    prime_cell = ws.cell(row=row, column=nb_cols)
    prime_cell.value = prime_sum
    prime_cell.number_format = '#,##0.00 €'
    return row + 1


def _write_total_row(ws, row, label, prime_sum, nb_cols):
    """Écrit une ligne de total général."""
    total_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=row, column=1, value=label)
    prime_cell = ws.cell(row=row, column=nb_cols)
    prime_cell.value = prime_sum
    prime_cell.number_format = '#,##0.00 €'


def generate_excel(all_data, output_path):
    """Génère le fichier Excel avec un onglet récap + un onglet par société."""
    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    colors = ["D6E4F0", "E2EFDA", "FCE4D6", "D9E2F3", "EDEDED", "FFF2CC", "D5E8D4", "F8CECC"]
    headers = ["Société", "Compagnie", "Objet du risque", "N° Police", "Échéance", "Périodicité", "Prime annuelle"]
    headers_soc = ["Compagnie", "Objet du risque", "N° Police", "Échéance", "Périodicité", "Prime annuelle"]
    col_widths = [30, 25, 45, 18, 12, 14, 18]
    col_widths_soc = [25, 45, 18, 12, 14, 18]

    # ========== ONGLET 1 : Récapitulatif global ==========
    ws = wb.active
    ws.title = "Récapitulatif"
    _write_headers(ws, headers)

    row = 2
    grand_total = 0.0

    for idx, (societe, contrats) in enumerate(all_data):
        fill_color = colors[idx % len(colors)]
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        sub_total = 0.0

        if not contrats:
            cell = ws.cell(row=row, column=1, value=societe)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", bold=True, size=10)
            for col in range(2, len(headers) + 1):
                c = ws.cell(row=row, column=col, value="")
                c.fill = row_fill
                c.border = thin_border
            row += 1
        else:
            for c in contrats:
                ws.cell(row=row, column=1, value=societe).fill = row_fill
                ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)
                ws.cell(row=row, column=2, value=c["compagnie"]).fill = row_fill
                ws.cell(row=row, column=3, value=c["objet_risque"]).fill = row_fill
                ws.cell(row=row, column=4, value=c["n_police"]).fill = row_fill
                ws.cell(row=row, column=5, value=c["echeance"]).fill = row_fill
                ws.cell(row=row, column=6, value=c["periodicite"]).fill = row_fill

                prime_cell = ws.cell(row=row, column=7)
                if c["prime_val"] is not None:
                    prime_cell.value = c["prime_val"]
                    prime_cell.number_format = '#,##0.00 €'
                    sub_total += c["prime_val"]
                else:
                    prime_cell.value = c["prime_str"] if c["prime_str"] else ""
                prime_cell.fill = row_fill

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
                row += 1

        # Sous-total par société
        row = _write_subtotal_row(ws, row, f"Sous-total {societe}", sub_total, len(headers))
        grand_total += sub_total

    # Total général
    _write_total_row(ws, row, "TOTAL GÉNÉRAL", grand_total, len(headers))

    _apply_styles_to_sheet(ws, headers, col_widths)
    ws.auto_filter.ref = f"A1:G{row}"

    # ========== ONGLETS PAR SOCIÉTÉ ==========
    for idx, (societe, contrats) in enumerate(all_data):
        # Nom d'onglet max 31 chars, sans caractères interdits
        sheet_name = re.sub(r"[\\/*?\[\]:]", "", societe)[:31]
        ws_soc = wb.create_sheet(title=sheet_name)

        # En-tête avec nom de société
        title_cell = ws_soc.cell(row=1, column=1, value=societe)
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
        ws_soc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_soc))

        # Ligne vide
        header_row = 3
        for col, h in enumerate(headers_soc, 1):
            cell = ws_soc.cell(row=header_row, column=col, value=h)
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        row_soc = header_row + 1
        sub_total = 0.0
        fill_color = colors[idx % len(colors)]
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for c in contrats:
            ws_soc.cell(row=row_soc, column=1, value=c["compagnie"]).fill = row_fill
            ws_soc.cell(row=row_soc, column=2, value=c["objet_risque"]).fill = row_fill
            ws_soc.cell(row=row_soc, column=3, value=c["n_police"]).fill = row_fill
            ws_soc.cell(row=row_soc, column=4, value=c["echeance"]).fill = row_fill
            ws_soc.cell(row=row_soc, column=5, value=c["periodicite"]).fill = row_fill

            prime_cell = ws_soc.cell(row=row_soc, column=6)
            if c["prime_val"] is not None:
                prime_cell.value = c["prime_val"]
                prime_cell.number_format = '#,##0.00 €'
                sub_total += c["prime_val"]
            else:
                prime_cell.value = c["prime_str"] if c["prime_str"] else ""
            prime_cell.fill = row_fill

            for col in range(1, len(headers_soc) + 1):
                ws_soc.cell(row=row_soc, column=col).border = thin_border
                ws_soc.cell(row=row_soc, column=col).alignment = Alignment(vertical="center", wrap_text=True)
            row_soc += 1

        # Sous-total
        _write_subtotal_row(ws_soc, row_soc, f"Total {societe}", sub_total, len(headers_soc))

        # Styles
        for i, w in enumerate(col_widths_soc, 1):
            ws_soc.column_dimensions[get_column_letter(i)].width = w
        ws_soc.freeze_panes = f"A{header_row + 1}"
        ws_soc.auto_filter.ref = f"A{header_row}:F{row_soc}"

    wb.save(output_path)
    return row - 2


def main():
    pdf_files = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
    print(f"Trouvé {len(pdf_files)} PDFs dans {PDF_DIR}\n")

    all_data = []
    total_contrats = 0

    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        print(f"Traitement: {filename}")

        raw_text = extract_text_from_pdf(pdf_path)
        societe = extract_societe(raw_text)
        if not societe:
            # Fallback: nom du fichier nettoyé
            societe = re.sub(r"^synth[eé]s[eé]\s+", "", filename.replace(".pdf", ""), flags=re.IGNORECASE).strip().upper()

        # Nettoyer le texte pour le parsing des contrats (supprime headers Brio)
        clean_text = clean_brio_headers(raw_text)
        contrats = extract_contrats(clean_text)
        print(f"  → Société: {societe} | {len(contrats)} contrat(s)")

        for c in contrats:
            prime_display = f"€ {c['prime_str']}" if c['prime_str'] else "(non renseignée)"
            print(f"    - {c['compagnie']} | {c['objet_risque'][:50]} | Éch: {c['echeance']} | Prime: {prime_display}")

        all_data.append((societe, contrats))
        total_contrats += len(contrats)

    print(f"\n{'='*60}")
    print(f"Total: {len(all_data)} sociétés, {total_contrats} contrats")
    print(f"Génération Excel: {OUTPUT_FILE}")

    nb_lignes = generate_excel(all_data, OUTPUT_FILE)
    print(f"Excel généré avec {nb_lignes} lignes de données.")
    print("Terminé !")


if __name__ == "__main__":
    main()
